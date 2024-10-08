from openpyxl import Workbook, load_workbook
from enum import Enum

class ColumnData:
    def __init__(self, mandatory: bool, name: str):
        self.mandatory = mandatory
        self.name = name

class Column(Enum):
    ID = ColumnData(True, "ID")
    SampleID = ColumnData(True, "Sample")
    Dataset = ColumnData(True, "Dataset")
    Tool = ColumnData(True, "Tool")
    Taxonomy = ColumnData(True, "Taxonomy")
    Profile = ColumnData(True, "Profile")
    ProfileColumns = ColumnData(True, "ProfileColumns")
    Gold = ColumnData(True, "Gold")
    GoldColumns = ColumnData(True, "GoldColumns")
    GoldStdTree = ColumnData(False, "GoldStdTree")
    AvailableSpecies = ColumnData(False, "AvailableSpecies")



def write_template_meta(output):
    wb = Workbook()
    act = wb.active

    act["A1"] = "Sample"
    act["B1"] = "Dataset"
    act["C1"] = "Tool"
    act["D1"] = "Profile"
    act["E1"] = "ProfileColumns"
    act["F1"] = "GoldStd"
    act["G1"] = "GoldStdColumns"
    
    wb.save(output)

class Meta:
    def __init__(self, path: str):
        self.path = path
        self.header = []
        self.header_dict = None
        self.workbook = load_workbook(path, data_only=True)
        self.load_header()
        self.validate()


    def validate(self):
        return all(e.value.name in self.header for e in Column if e.value.mandatory)
    

    def load_header(self):
        ac = self.workbook.active
        self.header = [e.value for e in ac[1]]
        self.header_dict = dict({(k, v)for v,k in enumerate(self.header)})
        
    def to_csv(self):
        print("Hello")

    def row_to_dict(self, row):
        return(dict(zip(self.header, [e.value for e in row])))
    
    def to_meta_dict(self):
        md = dict()
        for row in self.workbook.active:
            id = row[self.header_dict['ID']].value
            if id == None or id == "ID":
                continue
            md[id] = self.row_to_dict(row)
        return(md)
        
    